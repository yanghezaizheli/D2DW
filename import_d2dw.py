#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
D2DW インポートスクリプト (v0.1)
既存のGoogleスプレッドシート(エクスポートした .xlsx)を Supabase スキーマ用の seed.sql に変換します。

入力:
  --ledger  マスター台帳 (OTマップ座標リスト_*.xlsx)            … blocks(区域) と 集合住宅(建物) の座標・種別
  --block   区域スプシ (例: "01長住2[OT].xlsx") 複数指定可       … 各戸・号室・訪問ログ・ステータス
出力:
  --out     seed.sql (既定: seed.sql)。Supabase の SQL Editor に貼って実行します。
オプション:
  --dry-run SQLを書かず、抽出サマリ(件数・警告)だけ表示
  --base-date 移行データの基準日(不在回数→不在ログの日付に使用, 既定 2026-06-13)

設計上の判断(重要):
  - 表示値(前回訪問日時/前回面会日時/累計不在回数)は visits(訪問ログ)から集計する方針。
    既存シートの「会えた日付(直近/前回/前々回)」「不在回数」を初期ログに展開します。
  - 不在の個別日時は不明なため、不在回数ぶんの「不在ログ」を基準日でまとめて作成し、source='migration' で印を付けます。
  - 駐車場/空地/空家/空テナント等は除外。ステータス(訪問拒否/他言語/JW 等)はそのまま取り込みます。
  - 「訪問可能」かどうかは種別ごとの再訪間隔(visit_rules)で算出できるよう、ルール表も投入します。
  - 集合住宅(建物)→号室 は parent_id で表現。建物の座標・種別は台帳から取得します。

注意: 旧シートはレイアウトが不規則です。必ず --dry-run のサマリと seed.sql を目視確認してから実行してください。
"""
import argparse, re, sys, uuid, datetime

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install openpyxl", file=sys.stderr); sys.exit(1)

NS = uuid.UUID("d24d0000-0000-4000-8000-000000000001")  # 固定名前空間(決定的UUID用)

# ---------- 純粋関数(テスト済み) ----------
def parse_jp_date(s):
    """'26年6月13日' / '26/6/13' / '2026年6月13日' -> 'YYYY-MM-DD' or None"""
    if s is None: return None
    s = str(s).strip()
    if not s: return None
    m = re.match(r'^(\d{2,4})\D+(\d{1,2})\D+(\d{1,2})\D*$', s) or re.match(r'^(\d{2,4})/(\d{1,2})/(\d{1,2})$', s)
    if not m: return None
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100: y += 2000
    try: return datetime.date(y, mo, d).isoformat()
    except ValueError: return None

SKIP_PREFIX = '空'  # 空家/空地/空室/空テナント
SKIP_EXACT = {'駐車場', '', None}
def is_skip(name):
    n = (name or '').strip()
    return n in SKIP_EXACT or n.startswith(SKIP_PREFIX)

def is_building(name):
    return bool(re.search(r'(オートロック|エレベーター|階段)', name or ''))

STATUS_VOCAB = {'訪問可能', '訪問拒否', '他言語', 'JW', '不在', '転居', '長期不在', '空室'}
def looks_status(v):
    return (str(v).strip() in STATUS_VOCAB) if v is not None else False

def map_status(s):
    s = (s or '').strip()
    return s if s else '訪問可能'

def synth_visits(flag, absent_count, met_dates, base_date):
    """既存シートの状態を初期訪問ログに展開"""
    out = []
    for d in (met_dates or []):
        iso = parse_jp_date(d)
        if iso: out.append({'outcome': '会えた', 'at': iso})
    try: n = int(absent_count or 0)
    except (ValueError, TypeError): n = 0
    for _ in range(n):
        out.append({'outcome': '不在', 'at': base_date})
    if (flag or '').strip() == '投函':
        out.append({'outcome': '投函', 'at': base_date})
    return out

def det_uuid(*parts):
    return str(uuid.uuid5(NS, '|'.join(str(p) for p in parts)))

def sql_str(v):
    if v is None or v == '': return 'null'
    return "'" + str(v).replace("'", "''") + "'"

def parse_latlng(s):
    if not s: return (None, None)
    m = re.findall(r'(-?\d+\.\d+)', str(s))
    return (float(m[0]), float(m[1])) if len(m) >= 2 else (None, None)

# ---------- 台帳(マスター) ----------
LEDGER_HEADERS = ['エリア', 'type', '住所', '戸数', '緯度', '区域番号']
def load_ledger(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [[c.value for c in r] for r in ws.iter_rows()]
    # ヘッダ行を検出
    hi = next((i for i, r in enumerate(rows)
               if sum(1 for c in r if c and any(h in str(c) for h in ['エリア', 'type', '区域番号'])) >= 2), None)
    if hi is None:
        print("  [警告] 台帳のヘッダ行が見つかりません", file=sys.stderr); return {'blocks': [], 'buildings': []}
    hdr = [str(c).strip() if c else '' for c in rows[hi]]
    def col(*names):
        for i, h in enumerate(hdr):
            if any(n in h for n in names): return i
        return None
    ci = {k: col(*v) for k, v in {
        'area': ['エリア'], 'type': ['type', '種別'], 'name': ['建物名', '住所or'],
        'addr': ['住所'], 'units': ['戸数'], 'latlng': ['緯度', '経度'],
        'code': ['区域番号'], 'note': ['注記'], 'detail': ['詳細'],
    }.items()}
    # 'addr' が '住所or建物名' 列を誤検出した場合、厳密一致の '住所' 列に補正
    if ci['addr'] is not None and ci['addr'] == ci['name']:
        ci['addr'] = next((i for i, h in enumerate(hdr) if h.strip() == '住所'), None)
    blocks, buildings = [], []
    for r in rows[hi + 1:]:
        get = lambda key: (r[ci[key]] if ci[key] is not None and ci[key] < len(r) else None)
        name = get('name')
        if not name or not str(name).strip(): continue
        typ = (str(get('type')).strip() if get('type') else '')
        lat, lng = parse_latlng(get('latlng'))
        rec = {'name': str(name).strip(), 'type': typ, 'area': get('area'),
               'addr': get('addr'), 'code': get('code'), 'lat': lat, 'lng': lng,
               'units': get('units'), 'note': get('note')}
        if typ == 'LDR' or typ == '':
            blocks.append(rec)
        else:  # ST-M / EV-M / AL-M = 集合住宅(建物)
            buildings.append(rec)
    return {'blocks': blocks, 'buildings': buildings}

# ---------- 区域スプシ(各戸) ----------
def find_header(rows):
    for i, r in enumerate(rows):
        vals = [str(c).strip() if c else '' for c in r]
        if '号' in vals and '表記' in vals:
            return i, vals
    return None, None

def detect_context(rows):
    """タブ冒頭から 区域名 / 建物名 / 親区域(○○へ) を推定"""
    block_name = building = parent_area = None
    for r in rows[:8]:
        for v in r:
            if not v: continue
            s = str(v)
            if building is None and re.search(r'(オートロック|エレベーター|階段)', s) and re.search(r'[（(]', s):
                building = re.split(r'[（(]', s)[0].strip()
            if parent_area is None:
                m = re.search(r'(長住\d+-\d+)\s*へ', s)
                if m: parent_area = m.group(1)
            if block_name is None:
                m = re.search(r'(長住\d+-\d+[a-zA-Z]?)\s*[（(]', s)
                if m: block_name = m.group(1)
    return block_name, building, parent_area

def parse_block_workbook(path, base_date, warn):
    """各タブを走査して places(戸建て/集合住宅/号室) と visits, status を抽出"""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = [[c.value for c in r] for r in ws.iter_rows()]
        hi, hdr = find_header(rows)
        if hi is None:
            warn.append(f"[{sheet}] ヘッダ(号/表記)が見つからず skip"); continue
        block_name, tab_building, parent_area = detect_context(rows)
        if tab_building and not block_name:
            block_name = parent_area  # 建物タブは親区域に属する
        idx = {name: hdr.index(name) for name in hdr if name}
        c_go = idx.get('号'); c_name = idx.get('表記')
        c_flag = next((idx[k] for k in idx if 'フラグ' in k), None)
        c_abs = next((idx[k] for k in idx if '不在' in k and '回' in k), None)
        c_note = next((idx[k] for k in idx if '注記' in k), None)
        cur_building = tab_building  # 建物タブなら全行がこの建物の号室
        for r in rows[hi + 1:]:
            def cell(i): return r[i] if (i is not None and i < len(r)) else None
            label = cell(c_go); name = cell(c_name)
            label_s = str(label).strip() if label not in (None, '') else ''
            name_s = str(name).strip() if name else ''
            if not label_s and not name_s: continue
            if name_s == '表記' or label_s == '号': continue          # 繰り返しヘッダ
            if is_skip(name_s):                                       # 駐車場/空地/空家
                if not is_building(name_s): cur_building = None
                continue
            met = []
            for v in r:
                iso = parse_jp_date(v)
                if iso and iso not in met: met.append(iso)
            status = next((str(v).strip() for v in r if looks_status(v) and str(v).strip() != '不在'), '')
            if is_building(name_s):                                   # 建物(集合住宅)の行
                cur_building = re.split(r'[（(]', name_s)[0].strip()
                out.append({'tab': sheet, 'kind': '集合住宅', 'block_name': block_name,
                            'building': None, 'label': label_s, 'name': cur_building,
                            'flag': None, 'absent': None, 'met_dates': [], 'status': map_status(status), 'note': cell(c_note)})
                continue
            if not label_s and (cur_building or tab_building):         # 号室(号が空＝建物配下)
                kind, bld = '号室', (tab_building or cur_building)
            else:
                kind, bld = '戸建て', None
                cur_building = None
            out.append({'tab': sheet, 'kind': kind, 'block_name': block_name, 'building': bld,
                        'label': label_s, 'name': name_s, 'flag': cell(c_flag), 'absent': cell(c_abs),
                        'met_dates': met, 'status': map_status(status), 'note': cell(c_note)})
    return out

# ---------- seed.sql 生成 ----------
def build_sql(ledger, leaves, base_date):
    L = ["-- D2DW seed (自動生成) — 実行前に必ず目視確認してください", "begin;",
         "alter table visits add column if not exists source text not null default 'app';",
         "alter table blocks add column if not exists type text;",
         "create table if not exists visit_rules (type text primary key, absent_days int, flyer_days int, met_days int);"]
    for t, a, f, m in [('LDR', 30, 30, 90), ('ST-M', 90, 30, 180), ('EV-M', 90, 30, 180), ('AL-M', 90, 30, 180)]:
        L.append(f"insert into visit_rules values ('{t}',{a},{f},{m}) on conflict (type) do update set "
                 "absent_days=excluded.absent_days, flyer_days=excluded.flyer_days, met_days=excluded.met_days;")
    L.append("")

    # blocks(区域)
    block_id = {}
    for b in ledger['blocks']:
        bid = det_uuid('block', b['code'], b['name']); block_id[b['name']] = bid
        L.append("insert into blocks (id,code,name,area,type,lat,lng) values "
                 f"({sql_str(bid)},{sql_str(b['code'])},{sql_str(b['name'])},{sql_str(b['area'])},{sql_str(b['type'] or 'LDR')},"
                 f"{b['lat'] if b['lat'] is not None else 'null'},{b['lng'] if b['lng'] is not None else 'null'}) "
                 "on conflict (id) do update set area=excluded.area,type=excluded.type,lat=excluded.lat,lng=excluded.lng;")
    default_block = next(iter(block_id.values()), None)
    block_of = lambda nm: block_id.get(nm, default_block)

    # 建物(集合住宅): 台帳 ∪ 区域スプシで発見したもの。所属blockは区域スプシ文脈を優先。
    building_meta = {}   # name -> {addr,type,lat,lng,block_name}
    for bld in ledger['buildings']:
        building_meta[bld['name']] = {'addr': bld['addr'], 'type': bld['type'],
                                      'lat': bld['lat'], 'lng': bld['lng'], 'block_name': None}
    for p in leaves:
        if p['kind'] == '集合住宅':
            mm = building_meta.setdefault(p['name'], {'addr': None, 'type': None, 'lat': None, 'lng': None, 'block_name': None})
            if p['block_name']: mm['block_name'] = p['block_name']
        if p['kind'] == '号室' and p['building'] and p['block_name']:
            building_meta.setdefault(p['building'], {'addr': None, 'type': None, 'lat': None, 'lng': None, 'block_name': None})
            building_meta[p['building']]['block_name'] = building_meta[p['building']]['block_name'] or p['block_name']
    building_id = {}
    for name, mm in building_meta.items():
        pid = det_uuid('bld', name); building_id[name] = pid
        bidb = block_of(mm['block_name'])
        L.append("insert into places (id,block_id,kind,display_name,address,status,type,lat,lng) values "
                 f"({sql_str(pid)},{sql_str(bidb)},'集合住宅',{sql_str(name)},{sql_str(mm['addr'])},'訪問可能',{sql_str(mm['type'])},"
                 f"{mm['lat'] if mm['lat'] is not None else 'null'},{mm['lng'] if mm['lng'] is not None else 'null'}) "
                 "on conflict (id) do update set block_id=excluded.block_id,type=coalesce(excluded.type,places.type),"
                 "lat=coalesce(excluded.lat,places.lat),lng=coalesce(excluded.lng,places.lng);")
    L.append("")

    # 各戸・号室 + 訪問ログ
    nvisit = 0
    for p in leaves:
        if p['kind'] == '集合住宅': continue
        parent = sql_str(building_id.get(p['building'])) if p.get('building') else 'null'
        bid = block_of(p['block_name'])
        pid = det_uuid('place', p['tab'], p['kind'], p['label'], p['name'])
        L.append("insert into places (id,block_id,kind,parent_id,label,display_name,status,note) values "
                 f"({sql_str(pid)},{sql_str(bid)},{sql_str(p['kind'])},{parent},{sql_str(p['label'])},"
                 f"{sql_str(p['name'])},{sql_str(p['status'])},{sql_str(p['note'])}) "
                 "on conflict (id) do update set status=excluded.status,note=excluded.note,parent_id=excluded.parent_id;")
        for j, v in enumerate(synth_visits(p['flag'], p['absent'], p['met_dates'], base_date)):
            vid = det_uuid('visit', pid, v['outcome'], v['at'], j)
            L.append("insert into visits (id,place_id,outcome,visited_at,source) values "
                     f"({sql_str(vid)},{sql_str(pid)},{sql_str(v['outcome'])},{sql_str(v['at'])},'migration') on conflict (id) do nothing;")
            nvisit += 1
    L.append("commit;")
    return "\n".join(L), nvisit

# ---------- main ----------
def main():
    ap = argparse.ArgumentParser(description="D2DW: 旧スプシ(xlsx) → Supabase seed.sql")
    ap.add_argument('--ledger', help='マスター台帳 xlsx')
    ap.add_argument('--block', action='append', default=[], help='区域スプシ xlsx(複数可)')
    ap.add_argument('--out', default='seed.sql')
    ap.add_argument('--base-date', default='2026-06-13')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    warn = []
    ledger = {'blocks': [], 'buildings': []}
    if args.ledger:
        ledger = load_ledger(args.ledger)
    leaves = []
    for bp in args.block:
        leaves += parse_block_workbook(bp, args.base_date, warn)

    # サマリ
    from collections import Counter
    kc = Counter(p['kind'] for p in leaves)
    sc = Counter(p['status'] for p in leaves)
    print("=== 抽出サマリ ===")
    print(f"区域(blocks)     : {len(ledger['blocks'])}")
    print(f"建物(集合住宅)   : {len(ledger['buildings'])}")
    print(f"戸/号室(places)  : 戸建て={kc.get('戸建て',0)} 号室={kc.get('号室',0)} 建物参照={kc.get('集合住宅',0)}")
    print(f"ステータス内訳   : {dict(sc)}")
    if warn:
        print("--- 警告 ---")
        for w in warn[:40]: print(" -", w)

    if args.dry_run:
        print("\n(dry-run: SQLは出力していません)"); return

    sql, nv = build_sql(ledger, leaves, args.base_date)
    with open(args.out, 'w', encoding='utf-8') as f:
        f.write(sql)
    print(f"\n生成: {args.out} (訪問ログ {nv} 件)  ← 実行前に内容を確認してください")

if __name__ == '__main__':
    main()
